import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

start_date = datetime.date(2019, 1, 1)
end_date = datetime.date(2019, 1, 31)

for date_delta in range((end_date - start_date).days + 1):

    input_file = "input/sample.xlsx"
    output_file = "output/"+str(start_date+datetime.timedelta(date_delta))+"sample.xlsx"
    output_file_temp = "output/"+str(start_date+datetime.timedelta(date_delta))+"sample2.xlsx"

    wb = load_workbook(input_file)
    ws = wb['Sheet1']

    formula = ws['H2'].value
    print(ws.max_row)
    for row in range(2+1, ws.max_row+1):
        cell = 'H' + str(row)
        ws[cell] = Translator(formula, origin='H2').translate_formula(cell)
    wb.save(output_file)
    # wb.save(output_file2)

    # Load the input Excel file
    df = pd.read_excel(output_file, sheet_name='Sheet1',)

    # Drop unnecessary columns
    # df = df.drop(columns=['nouse', 'nouse.1', 'nouse.2'])

    # Pivot the DataFrame to reshape it
    print(df)
    pivot_df = df.pivot(index='formula', columns='item', values='price')
    pivot_df = pivot_df.sort_index()

    # Reset index and rename columns
    pivot_df = pivot_df.reset_index()
    pivot_df.columns.name = None

    # Save the output Excel file
    output_file = "output/"+str(start_date+datetime.timedelta(date_delta))+"sample.xlsx"
    pivot_df.to_excel(output_file, index=False)

print("Output generated successfully using pandas!")
