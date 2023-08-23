from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('sample.xlsx')
ws = wb['Sheet1']

formula = ws['D2'].value

for row in range(2+1, ws.max_row+1):
    cell = 'D' + str(row)
    ws[cell] = Translator(formula, origin='D2').translate_formula(cell)

wb.save('sample_res.xlsx')
